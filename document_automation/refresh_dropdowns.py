from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


CLIENT_NAME = input("Enter client name: ").strip()

client_folder = Path(CLIENT_NAME)
dump_folder = client_folder / "Dump"
excel_file = client_folder / "Document_Mapping.xlsx"

if not dump_folder.exists():
    raise FileNotFoundError(f"Dump folder not found: {dump_folder}")

if not excel_file.exists():
    raise FileNotFoundError(f"Excel file not found: {excel_file}")

files = [
    f.name
    for f in dump_folder.iterdir()
    if f.is_file()
]

wb = load_workbook(excel_file)

if "DropdownData" in wb.sheetnames:
    ws_dropdown = wb["DropdownData"]
    wb.remove(ws_dropdown)

ws_dropdown = wb.create_sheet("DropdownData")

for row_num, filename in enumerate(files, start=1):
    ws_dropdown.cell(row=row_num, column=1, value=filename)

ws_dropdown.sheet_state = "hidden"

ws_main = wb["Document Mapping"]

last_row = ws_main.max_row

dv = DataValidation(
    type="list",
    formula1=f"=DropdownData!$A$1:$A${len(files)}",
)

ws_main.add_data_validation(dv)

for row in range(2, last_row + 1):
    dv.add(ws_main[f"C{row}"])

wb.save(excel_file)

print(f"Added {len(files)} files to dropdown.")