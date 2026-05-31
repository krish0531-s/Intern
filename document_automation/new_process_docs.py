from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


def process_documents(client_folder):
    dump_folder = client_folder / "Dump"
    excel_file = client_folder / "Document_Mapping.xlsx"

    wb = load_workbook(excel_file)
    ws = wb["Document Mapping"]

    completed = 0
    pending = 0
    already_exists = 0
    not_found = 0

    for row in range(2, ws.max_row + 1):

        particulars = ws[f"A{row}"].value
        folder_name = ws[f"B{row}"].value
        selected_file = ws[f"C{row}"].value

        status_cell = ws[f"D{row}"]

        if not selected_file:
            status_cell.value = "Pending"
            pending += 1
            continue

        try:
            source_file = dump_folder / str(selected_file).strip()

            if not source_file.exists():
                status_cell.value = "File Not Found"
                not_found += 1
                continue

            extension = source_file.suffix

            destination_file = (
                client_folder
                / str(folder_name).strip()
                / f"{str(particulars).strip()}{extension}"
            )

            if destination_file.exists():
                status_cell.value = "Already Exists"
                already_exists += 1
                continue

            copy2(source_file, destination_file)

            status_cell.value = "Completed"
            completed += 1

        except Exception:
            status_cell.value = "Error"

    wb.save(excel_file)
    wb.close()

    print("\nSummary")
    print("-" * 30)
    print("Completed:", completed)
    print("Pending:", pending)
    print("Already Exists:", already_exists)
    print("File Not Found:", not_found)


def main():
    client_name = input(
        "Enter client name: "
    ).strip()

    if not client_name:
        print("Client name cannot be empty.")
        return

    client_folder = Path(client_name)

    if not client_folder.exists():
        print("Client folder not found.")
        return

    process_documents(client_folder)


if __name__ == "__main__":
    main()