from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


def main():
    client_name = input("Enter client name: ").strip()

    client_folder = Path(client_name)
    dump_folder = client_folder / "Dump"
    excel_file = client_folder / "Document_Mapping.xlsx"

    print("\n=== PATHS ===")
    print("Client Folder:", client_folder.resolve())
    print("Dump Folder:", dump_folder.resolve())
    print("Excel File:", excel_file.resolve())

    if not client_folder.exists():
        print("ERROR: Client folder not found.")
        return

    if not dump_folder.exists():
        print("ERROR: Dump folder not found.")
        return

    if not excel_file.exists():
        print("ERROR: Document_Mapping.xlsx not found.")
        return

    wb = load_workbook(excel_file)
    ws = wb["Document Mapping"]

    print("\n=== PROCESSING ROWS ===")

    for row in range(2, ws.max_row + 1):

        particulars = ws[f"A{row}"].value
        folder_name = ws[f"B{row}"].value
        selected_file = ws[f"C{row}"].value

        status_cell = ws[f"D{row}"]

        print("\n------------------------")
        print(f"Row: {row}")
        print(f"Particulars: {particulars}")
        print(f"Folder: {folder_name}")
        print(f"Selected File: {repr(selected_file)}")

        if not selected_file:
            status_cell.value = "Pending"
            print("Status: Pending (No file selected)")
            continue

        try:
            source_file = dump_folder / str(selected_file).strip()

            print("Source Path:", source_file)
            print("Source Exists:", source_file.exists())

            if not source_file.exists():
                status_cell.value = "File Not Found"
                print("Status: File Not Found")
                continue

            extension = source_file.suffix

            destination_file = (
                client_folder
                / str(folder_name).strip()
                / f"{str(particulars).strip()}{extension}"
            )

            print("Destination Path:", destination_file)

            if destination_file.exists():
                status_cell.value = "Already Exists"
                print("Status: Already Exists")
                continue

            copy2(source_file, destination_file)

            status_cell.value = "Completed"
            print("Status: Completed")

        except Exception as e:
            status_cell.value = f"Error"
            print("ERROR:", e)

    wb.save(excel_file)

    print("\n=== DONE ===")
    print("Workbook saved.")


if __name__ == "__main__":
    main()