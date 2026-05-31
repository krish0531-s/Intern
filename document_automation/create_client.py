from pathlib import Path
from openpyxl import Workbook


def read_checklist(checklist_file="checklist.txt"):
    documents = []
    folders = set()

    with open(checklist_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            document_type, folder_name = line.split("|")

            documents.append(
                {
                    "document_type": document_type.strip(),
                    "folder": folder_name.strip(),
                }
            )

            folders.add(folder_name.strip())

    return documents, folders


def create_excel(client_folder, documents):
    wb = Workbook()
    ws = wb.active

    ws.title = "Document Mapping"

    headers = [
        "Particulars",
        "Folder",
        "File From Dump",
        "Status",
    ]

    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, doc in enumerate(documents, start=2):
        ws.cell(row=row_num, column=1, value=doc["document_type"])
        ws.cell(row=row_num, column=2, value=doc["folder"])
        ws.cell(row=row_num, column=4, value="Pending")

    excel_path = client_folder / "Document_Mapping.xlsx"
    wb.save(excel_path)


def main():
    client_name = input("Enter client name: ").strip()

    if not client_name:
        print("Client name cannot be empty.")
        return

    documents, folders = read_checklist()

    client_folder = Path(client_name)
    client_folder.mkdir(exist_ok=True)

    (client_folder / "Dump").mkdir(exist_ok=True)

    for folder in folders:
        (client_folder / folder).mkdir(exist_ok=True)

    create_excel(client_folder, documents)

    print(f"\nCreated client structure: {client_folder}")


if __name__ == "__main__":
    main()