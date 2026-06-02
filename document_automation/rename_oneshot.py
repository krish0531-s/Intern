from pathlib import Path
from shutil import copy2
import subprocess
import platform
from tkinter import Tk, filedialog
import time
import os
import platform
import subprocess
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent


def read_checklist():
    documents = []
    folders = set()
    checklist_file = BASE_DIR / "checklist.txt"

    if not checklist_file.exists():
        print(
            f"ERROR: checklist.txt not found at:\n{checklist_file}"
        )
        return [], set()
    
    with open(checklist_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            document_type, folder_name = line.split("|")

            documents.append(
                {
                    "document_type": document_type.strip(),
                    "folder": folder_name.strip(),
                }
            )

            folders.add(folder_name.strip())

    return documents, folders


def select_source_folder():
    root = Tk()
    root.withdraw()

    folder = filedialog.askdirectory(
        title="Select Client Documents Folder"
    )

    root.destroy()

    return folder


def copy_files_to_dump(source_folder, dump_folder):
    copied = 0

    for file_path in Path(source_folder).rglob("*"):
        if not file_path.is_file():
            continue

        destination = dump_folder / file_path.name

        counter = 1

        while destination.exists():
            destination = (
                dump_folder
                / f"{file_path.stem}_{counter}{file_path.suffix}"
            )
            counter += 1

        copy2(file_path, destination)
        copied += 1

    return copied


def create_excel(client_folder, documents):
    wb = Workbook()
    ws = wb.active

    ws.title = "Document Mapping"

    headers = [
        "Particulars",
        "Folder",
        "File From Dump",
        "Name(Optional)",
        "Status",
    ]

    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, doc in enumerate(documents, start=2):
        ws.cell(row=row_num, column=1, value=doc["document_type"])
        ws.cell(row=row_num, column=2, value=doc["folder"])
        ws.cell(row=row_num, column=5, value="Pending")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[column_letter].width = max_length + 5

    excel_path = client_folder / "Document_Mapping.xlsx"
    wb.save(excel_path)
    wb.close()


def add_dropdowns(client_folder):
    dump_folder = client_folder / "Dump"
    excel_file = client_folder / "Document_Mapping.xlsx"

    files = [
        f.name
        for f in dump_folder.iterdir()
        if f.is_file()
    ]

    wb = load_workbook(excel_file)

    if "DropdownData" in wb.sheetnames:
        wb.remove(wb["DropdownData"])

    ws_dropdown = wb.create_sheet("DropdownData")

    for row_num, filename in enumerate(files, start=1):
        ws_dropdown.cell(
            row=row_num,
            column=1,
            value=filename
        )

    ws_dropdown.sheet_state = "hidden"

    ws_main = wb["Document Mapping"]

    if files:
        dv = DataValidation(
            type="list",
            formula1=f"=DropdownData!$A$1:$A${len(files)}"
        )

        ws_main.add_data_validation(dv)

        for row in range(2, ws_main.max_row + 1):
            dv.add(ws_main[f"C{row}"])

    wb.save(excel_file)
    wb.close()


# def open_excel(excel_file):
#     subprocess.Popen([
#         "open",
#         "-a",
#         "LibreOffice",
#         str(excel_file.resolve())
#     ])
def open_excel(excel_file):
    system = platform.system()

    if system == "Windows":
        os.startfile(str(excel_file.resolve()))

    elif system == "Darwin":
        subprocess.Popen([
            "open",
            "-a",
            "LibreOffice",
            str(excel_file.resolve())
        ])

    else:
        subprocess.Popen([
            "xdg-open",
            str(excel_file.resolve())
        ])



def process_documents(client_folder):
    dump_folder = client_folder / "Dump"
    excel_file = client_folder / "Document_Mapping.xlsx"

    wb = load_workbook(excel_file)
    ws = wb["Document Mapping"]

    completed = 0
    pending = 0
    already_exists = 0
    not_found = 0

    for row in range(2, ws.max_row + 1):

        particulars = ws[f"A{row}"].value
        folder_name = ws[f"B{row}"].value
        selected_file = ws[f"C{row}"].value
        custom_name= ws[f"D{row}"].value

        status_cell = ws[f"E{row}"]

        if not selected_file:
            status_cell.value = "Pending"
            pending += 1
            continue

        try:
            source_file = dump_folder / str(selected_file).strip()

            if not source_file.exists():
                status_cell.value = "File Not Found"
                not_found += 1
                continue

            extension = source_file.suffix

            particulars_clean= str(particulars).strip()
            if custom_name and particulars:
                final_name= (f"{str(custom_name).strip()}_"
                            f"{particulars_clean}")
            else:
                final_name = particulars_clean
            destination_file = (
                client_folder
                / str(folder_name).strip()
                / f"{str(final_name).strip()}{extension}"
            )

            if destination_file.exists():
                status_cell.value = "Already Exists"
                already_exists += 1
                continue

            copy2(source_file, destination_file)

            status_cell.value = "Completed"
            completed += 1

        except Exception:
            status_cell.value = "Error"

    wb.save(excel_file)
    wb.close()

    print("\nSummary")
    print("-" * 30)
    print("Completed:", completed)
    print("Pending:", pending)
    print("Already Exists:", already_exists)
    print("File Not Found:", not_found)


def main():
    client_name = input("Enter client name: ").strip()

    if not client_name:
        print("Client name cannot be empty.")
        return

    source_folder = select_source_folder()

    if not source_folder:
        print("No folder selected.")
        return

    documents, folders = read_checklist()

    client_folder = Path(client_name)
    client_folder.mkdir(exist_ok=True)

    dump_folder = client_folder / "Dump"
    dump_folder.mkdir(exist_ok=True)

    # for folder in folders:
    #     (client_folder / folder).mkdir(parents= True, exist_ok=True)
    #     source_folder,
    # copied = copy_files_to_dump(
    #     dump_folder
    # )

    for folder in folders:
        (client_folder / folder).mkdir(
            parents=True,
            exist_ok=True
        )

    copied = copy_files_to_dump(
        source_folder,
        dump_folder
    )

    print(f"Copied {copied} files.")

    create_excel(client_folder, documents)

    

    add_dropdowns(client_folder)

    excel_file = (
        client_folder
        / "Document_Mapping.xlsx"
    )

    time.sleep(1)

    open_excel(excel_file)

    print(
        "\nWorkbook opened."
    "\nComplete the mapping in Excel."
    "\nSave and CLOSE Excel."
    )

    print("\nWaiting for Excel to close...")

    while True:
        result = subprocess.run(
            ["tasklist"],
            capture_output=True,
            text=True
        )

        if "EXCEL.EXE" not in result.stdout.upper():
            break

        time.sleep(2)


    print("\nExcel closed.")
    print("Processing documents...")

    process_documents(client_folder)

    print("\nDone.")

if __name__ == "__main__":
    main()



